import torch,numpy as np,dgl,os,random
init_seed =0
dgl.seed(init_seed)
dgl.random.seed(init_seed)
torch.manual_seed(init_seed)
torch.cuda.manual_seed(init_seed)
torch.cuda.manual_seed_all(init_seed)
random.seed(init_seed)
np.random.seed(init_seed) # Random numbers used for numpy
torch.backends.cudnn.benchmark = False
os.environ['PYTHONHASHSEED'] = str(init_seed)  # Hash randomization is prohibited
os.environ['CUBLAS_WORKSPACE_CONFIG'] = ':4096:8'
#os.environ["CUDA_LAUNCH_BLOCKING"]='1.'#getMoreInformationUsed
torch.use_deterministic_algorithms(True) # Some operations use atomic operations, which are not deterministic algorithms, and are not guaranteed to be reproducible
torch.backends.cudnn.deterministic = True  # Make sure that the convolution algorithm returned each time is deterministical
torch.backends.cudnn.enabled = False  # disableCudnnToUseNonDeterministicAlgorithms

#Claim network and data
batch=100
inchannel=35
classes=2
epochs=200
lr=0.001
hidden_dim=64
nlayer=2
duplicatetime=1
samplenum=560
freenolandslide=False
trainscale=0.7
buffer=1000
seletcscale=1.0
runstep=[1,0,0,0]
weight=0.5
index=[i for i in range(inchannel)]
sub_o=[]
sub_a=[6,7,8,9,16,17,18,19,20,21,22,23,24,25,26,27,28,30]
sub_b=[6,7,8,9,17,18,19,20,21,22,23,24,25,26,27,28,29,30]
sub_c=[6,7,8,9,17,18,19,20,21,22,23,24,25,26,27,28,30]
sub_d=[2,10,26,30]
sub_e=[6,7,9,25]
sub_e1=[6,7,9]
sub_f=[2]
sub_h=[6,7,9,17,18,19,20,21,22,23,24,25,26,27,28,29,30]
sub_i=[6,7,9,17,18,19,20,21,22,23,24,25,26,27,28]
sub_j=[4,6,7,9,10,14,15,16,17,20,22,23,25,26,28,29,33]
sub_k=[4,6,9]
sub_l=[2,3,4,6,7,9]
sub_m=[2]
sub_n=[2,4,10]
sub_lmn=[2,3,4,6,7,9,10]
sub_p=[8,17,18,19,20,21,22,23,24,25,26,27,28,29,30]
sub_q=[7,16,17,23,29,31]
sub_r=[17,18,19,20,21,22,23,24,25,26,27,28]
sub_s=[17,18,19,20,21,22,23,24,25,26,27,28,29]
sub_t=[17,18,19,20,21,22,23,24,25,26,27,28,29,30]
selecti=list(set(index)-set(sub_lmn))
substr="sub_lmn"
activates=["sigmoid","elu","leaky_relu","relu"]
activef=activates[0]
from models import GNN_C,GNN_complex2,GNN_complex2std
from lstmgonv import crossgatconv,ginblock,gatblock,pnablock,gcnblock
import torch,os,pickle,dgl,numpy as np
from dgl.dataloading import GraphDataLoader
from sklearn.svm import SVC
from sklearn.model_selection import GridSearchCV
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier
from xgboost import XGBClassifier
from sklearn.neural_network import MLPClassifier as ANN
from sklearn.metrics import accuracy_score,precision_score,recall_score,roc_auc_score,f1_score,cohen_kappa_score,roc_curve
device=torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
print(f"inchannel:{len(selecti)}")
net=GNN_complex2(len(selecti),classes,crossgatconv,hiddensize=hidden_dim,p=0.0,acfun=activef).to(device)
clf=XGBClassifier(use_label_encoder=False,max_depth=5, learning_rate=0.1, n_estimators=100, objective='binary:logistic', nthread=-1, gamma=0, min_child_weight=1, max_delta_step=0, subsample=1, colsample_bytree=1, colsample_bylevel=1, reg_alpha=0, reg_lambda=1, scale_pos_weight=1, base_score=0.5, seed=0)
#clf=RandomForestClassifier(max_depth=5)
#clf=ANN()
#clf=LogisticRegression()
#clf=SVC(kernel="linear",probability=True)


#
# tifpath="D:/nepal/newdata/dataset/LS_data_mergeadd.tif"
# excelpath="D:/nepal/newdata/dem/2007/slopunit/selected_sample.xlsx"
# possheet,negsheet=1,2
# traindataset=LSDataset(tifpath,excelpath,possheet,negsheet)
# possheet2,negsheet2=3,4
# testdataset=LSDataset(tifpath,excelpath,possheet2,negsheet2)
# traindataloader=GraphDataLoader(traindataset,batch_size=batch,drop_last=True)
# testdataloader=GraphDataLoader(testdataset,batch_size=batch,drop_last=True)
prename="freenolandslide" if freenolandslide is True else ""
trainscale=str(trainscale) if trainscale!=0.5 else ""
prefex="buffer{}".format(buffer) if buffer!=1000 else ""
# print('../dataset/'+ "{}{}randomtest_samples{}.pkl".format(prefex,prename,trainscale));exit(0)
with open('dataset/'+ "{}{}randomtest_samples{}.pkl".format(prefex,prename,trainscale), 'rb') as f:
    testpacks = pickle.load(f)
    # if samplenum==1000 and buffer==1000:
    #     testpacks=testpacks[int(len(testpacks)*0.4):]
    onefeats,retangelfeats,gs,labels,slopunits=zip(*testpacks)#todo:It is best not to mix dglgraph with pickle
    batch_testgs=[];batch_testonefeats=[]
    batch_testlabels=[]
    print("testsample_len:",len(testpacks))
    for i in range(len(onefeats)//batch):
        batch_testgs.append(dgl.batch(gs[i:i+batch]));batch_testonefeats.append(torch.cat(onefeats[i:i+batch],dim=0))
        batch_testlabels.append(torch.cat(labels[i:i+batch],dim=0))
    test_xs, testlabel = torch.cat(onefeats, dim=0).numpy(), torch.cat(labels, dim=0).flatten().numpy()
with open('dataset/'+ "{}{}randomtrain_samples{}.pkl".format(prefex,prename,trainscale), 'rb') as f:
    trainpacks = pickle.load(f)
    random.shuffle(trainpacks)
    # if samplenum==1000 and buffer==1000:
    #     with open('../dataset/LSM/{}/'.format(samplenum) + "{}{}randomtest_samples{}.pkl".format(prefex, prename,
    #                                                                                              trainscale),
    #               'rb') as f:
    #         testpacks = pickle.load(f)
    #         addpacks=testpacks[:int(len(testpacks)*0.4)]
    #trainpacks=samplebyscale(trainpacks,seletcscale)
    onefeats, retangelfeats, gs, labels, slopunits = zip(*trainpacks)
    batch_traings = [];batch_trainonefeats=[]
    batch_trainlabels = []
    print("trainsample_len:", len(trainpacks))
    for i in range(len(onefeats) // batch):
        batch_traings.append(dgl.batch(gs[i:i + batch]));batch_trainonefeats.append(torch.cat(onefeats[i:i+batch],dim=0))
        batch_trainlabels.append(torch.cat(labels[i:i+ batch], dim=0))
    train_xs,trainlabel=torch.cat(onefeats,dim=0).numpy(),torch.cat(labels,dim=0).flatten().numpy()
from sampleselect import pkldataset
testdataset=pkldataset('dataset/'+ "{}{}randomtest_samples{}.pkl".format(prefex,prename,trainscale))
testdataloader=GraphDataLoader(testdataset,batch_size=batch,shuffle=False,drop_last=False)
traindataset=pkldataset('dataset/'+ "{}{}randomtrain_samples{}.pkl".format(prefex,prename,trainscale))
traindataloader=GraphDataLoader(traindataset,batch_size=batch,shuffle=True,drop_last=False)
#Training, evaluation, etc.
import torch.optim as optim,time
import torch.nn as nn
import torch.nn.functional as F
criterion=nn.CrossEntropyLoss()
#criterion=CrossEntropyLoss(e=0.04,nclass=classes,newsimplemode=3)
#criterion=LCMloss(classes,hidden_size=hidden_dim).to(device)
# log2File
activename="" if activef=="sigmoid" else activef
duplicatename="" if duplicatetime==1 else duplicatetime
training_name="lsm_{}{}{}{}{}_epoch{}batch{}lr{}hidden{}nlayer{}(2)_{}_".format(f"crossgatcomplex3{activename}{duplicatename}_trainbaranchXGBoost0.5_",prefex,prename,samplenum,trainscale,epochs,batch,lr,hidden_dim,nlayer,substr)
metricpath = "trained_model/metric/" + training_name + "/"
modelpath = "trained_model/model/" + training_name + "/"
if os.path.exists(metricpath) is False:
    os.makedirs(metricpath)
if os.path.exists(modelpath) is False:
    os.makedirs(modelpath)
if runstep[0] == 1:
    import copy
    argtmp=[[i,copy.deepcopy(net),clf] for i in range(duplicatetime)]
    def mapfun(argtuple):
        sec,net,clf=argtuple
        import time,numpy as np
        time.sleep(sec)
        clf.fit(train_xs[:, selecti], trainlabel)
        # Save model
        with open(modelpath + training_name + "{}.pkl".format(sec), 'wb') as f:
            pickle.dump(clf, f)
        print("ok")
        optimizer = optim.Adam(net.parameters(), lr=lr, eps=1e-06, weight_decay=0.001)
        lambda1 = lambda iter: (1 - iter / epochs) ** 0.9
        lr_scheduler = torch.optim.lr_scheduler.LambdaLR(optimizer, lr_lambda=lambda1)  # PolyStrategy
        #strics_file=open(metricpath+training_name+"strics_global.txt",'a+',buffering=1)#Recording accuracy
        #strics_file.write("mean_train_loss"+"    "+"mean_train_acc"+"    "+"mean_test_loss"+"    "+"mean_testacc"+"    "+"mean_p"+"    "+"mean_recall"+"    "+"mean_auc"+"    "+"mean_f1"+"    "+"mean_kappa"+"\n")
        columns=["mean_train_loss","mean_train_acc","mean_test_loss","mean_testacc","mean_branchacc","mean_p","mean_recall","mean_auc","mean_f1","mean_kappa"]
        strics=[];curves=[];rocshape=None
        #iterative training
        bestkappa=0.;bestmodel=None
        bestc_epoch=0.;bestacc,bestp,bestrecall,bestf1,bestauc=0,0,0,0,0
        runningtime=0.
        for i in range(epochs):
            runningloss=0.
            runningacc=0.
            #for _,_,batchg,labels,_ in traindataloader:
            #for onefeat,batchg,labels in zip(batch_trainonefeats,batch_traings,batch_trainlabels):
            for onefeat, rectanglefeat, batchg, labels, slopunit in traindataloader:
                batchg=batchg.to(device);onefeat=onefeat.to(device)
                labels=labels.flatten().to(device)
                optimizer.zero_grad()
                pred=net(batchg,batchg.ndata["feat"][:,selecti])#batch,classes
                #pred2, feature2 = net(batchg, RandomAugmentation(batchg.ndata["feat"][:, selecti]))
                loss=criterion(pred,labels)#+0.1*criterion(pred2,pred.argmax(dim=-1).detach())
                loss.backward()
                optimizer.step()
                acc=accuracy_score(labels.cpu().numpy(),pred.argmax(dim=-1).detach().cpu().numpy())
                runningloss=runningloss+loss.item()
                runningacc=runningacc+acc
            lr_scheduler.step()
            meantrainloss=runningloss/traindataloader.__len__()#len(batch_traings)#
            meantrainacc=runningacc/traindataloader.__len__()#len(batch_traings)#
            print("meantrainloss:",meantrainloss,"meantrainacc:",meantrainacc)
            #Test
            net.eval()
            runningloss,runningacc,runningp,runningrecall,runningkappa,runningf1,runningroc=0,0,0,0,0,0,0
            time1 = time.time()
            tmprocs=[]
            runningbranchacc=0.
            epoch=0
            #for _,_,batchg,labels,_ in testdataloader:
            for batchid,onefeat,batchg,labels in zip([i for i in range(len(testpacks))],batch_testonefeats,batch_testgs,batch_testlabels):
            #for onefeat, rectanglefeat, batchg, labels, slopunit in testdataloader:
                batchg=batchg.to(device);onefeat=onefeat.to(device)
                labels=labels.flatten().to(device)
                with torch.no_grad():
                    pred=net(batchg,batchg.ndata["feat"][:,selecti])#batch,classes
                    pred2=(1-weight)*torch.from_numpy(clf.predict_proba(onefeat[:,selecti].cpu().numpy())).to(device)+weight*pred
                loss=criterion(pred,labels)
                acc=accuracy_score(labels.cpu().numpy(),pred.argmax(dim=-1).detach().cpu().numpy())
                acc2= accuracy_score(labels.cpu().numpy(), pred2.argmax(dim=-1).detach().cpu().numpy())
                precision=precision_score(labels.cpu().numpy(), pred2.argmax(dim=-1).detach().cpu().numpy())
                recall=recall_score(labels.cpu().numpy(), pred2.argmax(dim=-1).detach().cpu().numpy())
                roc=roc_auc_score(labels.cpu().numpy(), pred2[:,1].detach().cpu().numpy())
                f1=f1_score(labels.cpu().numpy(),pred2.argmax(dim=-1).detach().cpu().numpy())
                kappa= cohen_kappa_score(labels.cpu().numpy(), pred2.argmax(dim=-1).detach().cpu().numpy())
                curvetuple = roc_curve(labels.cpu().numpy(), pred2[:, 1].detach().cpu().numpy())
                curve = np.stack(curvetuple, axis=-1)
                tmprocs.append(curve)

                runningloss=runningloss+loss.item()
                runningacc=runningacc+acc
                runningp=runningp+precision
                runningrecall=runningrecall+recall
                runningroc=runningroc+roc
                runningkappa=runningkappa+kappa
                runningf1=runningf1+f1
                runningbranchacc=runningbranchacc+acc2
            time2 = time.time()
            runningtime = runningtime+time2 - time1
            meantestloss=runningloss/len(batch_testgs)#testdataloader.__len__()#
            meantestacc= runningacc/len(batch_testgs)#testdataloader.__len__()#
            meanbranchacc=runningbranchacc/len(batch_testgs)#testdataloader.__len__()#
            meantestp= runningp/len(batch_testgs)#testdataloader.__len__()#
            meantestrecall= runningrecall /len(batch_testgs)#testdataloader.__len__()#
            meantestroc= runningroc/len(batch_testgs)#testdataloader.__len__()#
            meantestf1= runningf1/len(batch_testgs)#testdataloader.__len__()#
            meantestkappa= runningkappa/len(batch_testgs)#testdataloader.__len__()#
            meancurve=np.concatenate(tmprocs,axis=0)

            print("meantestloss:",meantestloss,"meantestacc:",meantestacc,"meantestkappa:",meantestkappa)
            # strics_file.write(str(meantrainloss)+"     "+str(meantrainacc)+"     "+str(meantestloss)+"     "+str(meantestacc)+
            #                   "     "+str(meantestp)+"     "+str(meantestrecall)+"     "+str(meantestroc)+"     "+str(meantestf1)
            #                   +"     "+str(meantestkappa)+"\n")
            tmp=[meantrainloss,meantrainacc,meantestloss,meantestacc,meanbranchacc,meantestp,meantestrecall,meantestroc,meantestf1,meantestkappa]
            if len(tmp)==len(columns):
                strics.append(tmp)
                if i==0:
                    rocshape=list(meancurve.shape)
                meancurve=F.interpolate(torch.tensor(meancurve).unsqueeze(dim=0).unsqueeze(dim=0),size=rocshape).squeeze().numpy()
                curves.append(meancurve)
            else:
                print("record strics length is not enough");exit(0)
            if bestkappa<=meantestkappa:
                bestkappa=meantestkappa
                bestc_epoch=i
                bestmodel=net
                bestacc=meantestacc;bestauc=meantestroc
                torch.save(bestmodel.state_dict(), modelpath + training_name + "{}.ph".format(bestc_epoch))
        runningtime=runningtime/epochs
        import numpy as np, pandas as pd, glob
        columns2 = ["bestepoch", "bestacc", "bestauc", "bestkappa","runningtime"]
        data2 = np.array([[bestc_epoch, bestacc, bestauc, bestkappa,runningtime]])
        df2 = pd.DataFrame(data2, columns=columns2)
        data = np.array(strics)
        df = pd.DataFrame(data, columns=columns)
        columns3=sum([["fpr{}".format(i),"tpr{}".format(i),"thres{}".format(i)] for i in range(epochs)],[])#sum(listlist,[])变为单列表
        df3=pd.DataFrame(np.concatenate(curves,axis=-1),columns=columns3)
        # strics_file.write("bestepoch:"+str(bestc_epoch)+"   bestacc:"+str(bestacc)+"    bestauc:"+str(bestauc)+"    bestkappa:"+str(bestkappa)+"\n")
        # strics_file.close()
        print("bestepoch:"+str(bestc_epoch)+"   bestacc:"+str(bestacc)+"    bestauc:"+str(bestauc)+"    bestkappa:"+str(bestkappa))
        return df,df2,df3,bestmodel,bestc_epoch
    from multiprocessing.dummy import Pool
    pool=Pool()
    result=list(map(mapfun,argtmp))
    pool.close()
    # result=[]
    # for net in argtmp:
    #     result.append(mapfun(net))
    import pandas as pd, glob
    stricsfullpath = metricpath + "strics_global.xlsx"
    i=0;tmps=[]
    for df,df2,df3,bestmodel,bestepoch in result:
        if (i==0) and os.path.exists(stricsfullpath) is False:
            writer = pd.ExcelWriter(stricsfullpath, engine="openpyxl", mode="a+")
            df.to_excel(writer, sheet_name="all")
            df2.to_excel(writer, sheet_name="best")
            df3.to_excel(writer, sheet_name="roc")
            writer.save()
            writer.close()
        else:
            from openpyxl import load_workbook
            book = load_workbook(stricsfullpath)
            writer = pd.ExcelWriter(stricsfullpath, engine="openpyxl", mode="a+")
            writer.book = book
            tmpdf = pd.read_excel(stricsfullpath, sheet_name=None)  # None means taking the left and right worksheets
            sheetnum = len(tmpdf.keys())
            df.to_excel(writer, sheet_name="all")
            df2.to_excel(writer, sheet_name="best")
            df3.to_excel(writer, sheet_name="roc")
            writer.save()
            writer.close()
            print("add")
        tmps.append(df.iloc[bestepoch].values)
        torch.save(bestmodel.state_dict(), modelpath + training_name + "{}{}.ph".format(bestepoch,i))
        i=i+1
    bestrow=np.stack(tmps,axis=0)
    bestmean=bestrow.mean(axis=0)
    beststd=bestrow.std(axis=0)
    bestdata=np.stack([bestmean,beststd],axis=0)
    bestdf=pd.DataFrame(bestdata,columns=result[0][0].columns)
    from openpyxl import load_workbook
    book = load_workbook(stricsfullpath)
    writer = pd.ExcelWriter(stricsfullpath, engine="openpyxl", mode="a+")
    writer.book = book
    tmpdf = pd.read_excel(stricsfullpath, sheet_name=None)  #None means taking the left and right worksheets
    sheetnum = len(tmpdf.keys())
    bestdf.to_excel(writer,sheet_name="bestmean")
    writer.save()
    writer.close()
    print("bestmean:",bestmean)
#Big picture test
if runstep[1]==1:
    bestc_epoch=3;sec=0
    net.load_state_dict(torch.load(modelpath+training_name+str(bestc_epoch)+".ph"))
    net.eval()
    with open(modelpath+training_name+"{}.pkl".format(sec), 'rb') as f:
        clf=pickle.load(f)
    # Save in dataset
    test2imagepath = "trained_model/testtoimage/" + training_name + "/"
    if os.path.exists(test2imagepath) is False:
        os.makedirs(test2imagepath)
    #Load data
    import glob,pickle,numpy as np,pandas as pd
    packroot="E:/tmpallsamples/"
    packpaths=glob.glob(packroot+"*.pkl")
    #datas=[]
    argtmp=[[i,packpath] for i,packpath in enumerate(packpaths)]
    def mapfun(argtuple):
        i,packpath=argtuple
        with open(packpath,"rb") as f:
            onefeat,rectanglefeat,gs,labels,slopunits=pickle.load(f)
        print(gs,gs.batch_size)
        gs=gs.to(device)
        with torch.no_grad():
            pred=net(gs,gs.ndata["feat"][:,selecti])
            pred2 = (1 - weight) * torch.from_numpy(
                clf.predict_proba(onefeat[:, selecti].cpu().numpy())).to(device) + weight * pred
        data=torch.cat([slopunits,pred2[:,1:2].cpu()],dim=-1).numpy()#m,1 m,2
        print("index{}".format(i),data.shape)
        return data
    from multiprocessing.dummy import Pool
    pool=Pool(processes=3)
    datas=list(pool.map(mapfun,argtmp))
    pool.close()
    # for argtuple in argtmp:
    #     datas.append(mapfun(argtuple))

    datas=np.concatenate(datas,axis=0)#,2
    tifpath = "*.tif"
    from sampleselect import LSDataset
    excelpath = r"template/template.xlsx"
    possheet, negsheet = 1, 2
    dataset = LSDataset(tifpath, excelpath)
    dataset.OutLinkSU(datas, test2imagepath + training_name + "result.tif")
if runstep[2]==1:#explanation for CGBoost
    import pandas as pd
    bestc_epoch = 3;sec = 0
    net.load_state_dict(torch.load(modelpath + training_name + str(bestc_epoch) + ".ph"))
    with open(modelpath+training_name+"{}.pkl".format(sec), 'rb') as f:
        clf=pickle.load(f)
    print("adaptiveWeights:",net.weight)
    #TWO KINDS OF DATA
    "Test set full data"
    select_num=560;possheet= 1;negsheet= 3;freenolandslide= False;trainscale= 0.7;buffer= 1000;
    othertest= 0;otherposfix="explainer"
    prename = "freenolandslide" if freenolandslide is True else ""
    trainscale = str(trainscale) if trainscale != 0.5 else ""
    prefex = "buffer{}".format(buffer) if buffer != 1000 else ""
    otherprefex = "other" if othertest == 1 else ""
    if possheet == 0 and negsheet == 2:
        dname = "train"
    elif othertest == 1:
        dname = "test2"
    elif othertest == 2:
        dname = "pos"
    else:
        dname = "test"
    with open('dataset/'.format(select_num)+"{}{}random{}_samples{}{}.pkl".format(prefex,prename,dname,trainscale,otherposfix), 'rb') as f:
        testpacks= pickle.load(f)
    explainindex=[18,31,96,124,145,157,169,309,326,330]
    onefeats, rectanglefeats, gss, labels, slopunits=zip(*testpacks)
    print(labels[0].shape)
    onefeats=torch.cat(onefeats,dim=0)
    columns = ["Elevation", "Slope angle", "Slope aspect", "Profile curvature", "Plan curvature", "Relative relief", "TWI", "SPI", "TRI",
               "Newmark displacement", "Lithology", "Distance to faults", "Soil type", "Land use", "Distance to roads", "Distance to rivers", "Average annual rainfall",
               "B1", "B2", "B3", "B4", "B5", "B6", "B7", "B8", "B9", "B10", "B11", "B12", "CRMR",
               "CAMR", "MI", "PGA", "PGV", "Epicenter distance", "label"]
    from functions import listbyindice
    df = pd.DataFrame(onefeats[:, selecti].numpy(), columns=listbyindice(columns, selecti))
    print(df.shape)
    "Typical 10 sample graph data"
    typicalsampleg=dgl.batch(listbyindice(gss,explainindex)).to(device)
    relativenid=[]
    for tmpggg in listbyindice(gss,explainindex):
        relativenid.append(tmpggg.nodes().reshape(-1,1))
    realnid=torch.arange(typicalsampleg.num_nodes()).reshape(-1,1)
    nidtransform = torch.cat([realnid, torch.cat(relativenid, dim=0)], dim=1)
    pd.DataFrame(nidtransform.numpy(), columns=["realnid", "relativenid"]).to_excel(
        "tmpdir/nidtransform.xls")
    print(typicalsampleg.batch_num_nodes())
    print(listbyindice(slopunits,explainindex))
    print(listbyindice(labels,explainindex))


    #xgBoostExplained
    import shap,copy
    from _waterfall import waterfall
    xgbexplainer=shap.Explainer(clf)
    shapvalues =xgbexplainer(df)
    newshap=copy.deepcopy(shapvalues)
    print(shapvalues.clustering)
    print(shapvalues.base_values.shape)#(n,)
    print(shapvalues.values.shape)#shapvalue？？
    print(shapvalues.data.shape)#feat
    print(shapvalues.feature_names)
    print(shapvalues.lower_bounds)
    print(shapvalues.upper_bounds)
    print(shapvalues.display_data)

    #Merge prcp time series
    prcp_sumvalues=np.mean(newshap.values[:,10:22],axis=1,keepdims=True)
    newshap.values=np.concatenate([newshap.values[:,:10],prcp_sumvalues,newshap.values[:,22:]],axis=1)
    prcp_sumdata = np.mean(newshap.data[:, 10:22], axis=1, keepdims=True)
    newshap.data = np.concatenate([newshap.data[:, :10], prcp_sumdata, newshap.data[:,22:]], axis=1)
    newshap.feature_names=newshap.feature_names[:10]+["MRS"]+newshap.feature_names[22:]
    #保存new shapdf
    newshapdf = pd.DataFrame(newshap.values, columns=newshap.feature_names)
    newshapdf.to_excel("tmpdir/xgboostnewshap.xls")

    shap.plots.beeswarm(newshap,max_display=35)#summary plot
    shap.summary_plot(newshap, newshapdf, plot_type="bar", max_display=100)  # shap average importance plot
    "Typical sample explanation"
    waterfall(newshap[18],max_display=100,fontsizeadd=3,equal_color="#800080")#Single sample feature contribution map, the value after ylabel is the feature value
    waterfall(newshap[31], max_display=100,fontsizeadd=3,equal_color="#800080")  # Single sample feature contribution map
    #waterfall(newshap[96], max_display=100,fontsizeadd=3,equal_color="#800080")  # Single sample feature contribution map
    #waterfall(newshap[124], max_display=100,fontsizeadd=3,equal_color="#800080")  # Single sample feature contribution map
    waterfall(newshap[145], max_display=100,fontsizeadd=3,equal_color="#800080")  # Single sample feature contribution map
    #waterfall(newshap[157], max_display=100,fontsizeadd=3,equal_color="#800080")  # Single sample feature contribution map
    waterfall(newshap[169], max_display=100,fontsizeadd=3,equal_color="#800080")  # Single sample feature contribution map
    waterfall(newshap[309], max_display=100,fontsizeadd=3,equal_color="#800080")  # Single sample feature contribution map
    waterfall(newshap[326], max_display=100,fontsizeadd=3,equal_color="#800080")  # Single sample feature contribution map
    waterfall(newshap[330], max_display=100,fontsizeadd=3,equal_color="#800080")  # Single sample feature contribution map
    # shap.plots.bar(shapvalues)#Bar chart
    # shap.plots.scatter(shapvalues[:,"slop"])#Contribution distribution of single feature
    shapdf = pd.DataFrame(shapvalues.values, columns=shapvalues.feature_names)
    shapdf.to_excel("tmpdir/xgboostshap.xls")

    # forecast result
    class GNN_complex2_result(nn.Module):
        def __init__(self):
            super(GNN_complex2_result, self).__init__()
            self.nlayer = net.nlayer
            # relation
            self.conv0=net.conv0
            for layer in range(net.nlayer)[1:-1]:
                setattr(self, "conv{}".format(layer), getattr(net,"conv{}".format(layer)))
            setattr(self, "conv{}".format(net.nlayer - 1), getattr(net,"conv{}".format(net.nlayer - 1)))
            #node
            setattr(self, "conv{}_1".format(0),getattr(net,"conv{}_1".format(0)))
            for layer in range(net.nlayer)[1:-1]:
                setattr(self, "conv{}_1".format(layer),getattr(net,"conv{}_1".format(layer)))
            setattr(self, "conv{}_1".format(net.nlayer - 1),getattr(net,"conv{}_1".format(net.nlayer - 1)))
            self.elu = net.elu
            self.modenode = net.modenode
            self.dropout = net.dropout
            self.weight =net.weight

        def layer_normal(self, x, dim=None):
            if dim is None:
                std, mean = torch.std_mean(x)
            else:
                std, mean = torch.std_mean(x, dim, keepdim=True)
            return (x - mean) / (std + 1e-05)

        def forward(self, graph, feat,eweight=None):  # Normal graph classification and terminal node classification by adding terminal nodes
            g,ndata=graph,feat
            with g.local_scope():
                # node
                g.ndata["feat"] = ndata
                onefeat = dgl.readout_nodes(g, "feat", op="mean")
                tmpg = dgl.knn_graph(onefeat, len(onefeat))  # Create sample diagram
                h1 = self.elu(self.layer_normal(getattr(self, "conv{}_1".format(0))(tmpg, onefeat)))
                h1 = self.dropout(h1)
                # realtion
                h = self.elu(self.layer_normal(getattr(self, "conv{}".format(0))(g, ndata)))
                h = self.dropout(h)
                g.ndata["tmp"] = h
                readout = dgl.readout_nodes(g, "tmp", op="max")
                for index, layer in enumerate(range(self.nlayer)[1:-1]):
                    # node
                    h1 = self.elu(self.layer_normal(getattr(self, "conv{}_1".format(layer))(tmpg, h1 + readout)))
                    h1 = self.dropout(h1)
                    # relation
                    h = self.elu(self.layer_normal(getattr(self, "conv{}".format(layer))(g, h)))
                    h = self.dropout(h)
                    g.ndata["tmp"] = h
                    readout = dgl.readout_nodes(g, "tmp", op="max")
                # node
                noderesult = getattr(self, "conv{}_1".format(self.nlayer - 1))(tmpg, h1 + readout)
                # relation
                h = getattr(self, "conv{}".format(self.nlayer - 1))(g, h)
                g.ndata["h"] = h
                relationresult = dgl.readout_nodes(g, "h", op="max")
                h = self.weight *relationresult+ (1 - self.weight) * noderesult
                h = torch.softmax(h, dim=-1)
            return h,relationresult[:,1],noderesult[:,1]
    onefeat= onefeats[explainindex]#Onefeats has been turned into tensor before
    pred,relationresult,noderesult = GNN_complex2_result()(typicalsampleg,typicalsampleg.ndata["feat"][:, selecti])
    pred_=torch.from_numpy(clf.predict_proba(onefeat[:, selecti].cpu().numpy())).to(device)
    pred2 = (1 - weight) *pred_ + weight * pred
    tmpdata=torch.stack([relationresult,noderesult,pred_[:,1],pred2[:,1]],dim=1).detach().cpu().numpy()
    pd.DataFrame(tmpdata,columns=["realtionresult","noderesult","xgboostresult","zongresult"]).to_excel("tmpdir/predict.xls")
    print(relationresult,noderesult,pred_[:,1],pred2[:,1])
    # GNN explained
    class GNN_complex2_explaing(nn.Module):
        def __init__(self):
            super(GNN_complex2_explaing, self).__init__()
            self.nlayer = net.nlayer
            # relation
            self.conv0=net.conv0
            for layer in range(net.nlayer)[1:-1]:
                setattr(self, "conv{}".format(layer), getattr(net,"conv{}".format(layer)))
            setattr(self, "conv{}".format(net.nlayer - 1), getattr(net,"conv{}".format(net.nlayer - 1)))
            #node
            setattr(self, "conv{}_1".format(0),getattr(net,"conv{}_1".format(0)))
            for layer in range(net.nlayer)[1:-1]:
                setattr(self, "conv{}_1".format(layer),getattr(net,"conv{}_1".format(layer)))
            setattr(self, "conv{}_1".format(net.nlayer - 1),getattr(net,"conv{}_1".format(net.nlayer - 1)))
            self.elu = net.elu
            self.modenode = net.modenode
            self.dropout = net.dropout
            self.weight =net.weight

        def layer_normal(self, x, dim=None):
            if dim is None:
                std, mean = torch.std_mean(x)
            else:
                std, mean = torch.std_mean(x, dim, keepdim=True)
            return (x - mean) / (std + 1e-05)

        def forward(self, graph, feat,eweight=None):  # Normal graph classification and terminal node classification by adding terminal nodes
            g,ndata=graph,feat
            with g.local_scope():
                # node
                g.ndata["feat"] = ndata
                onefeat = dgl.readout_nodes(g, "feat", op="mean")
                tmpg = dgl.knn_graph(onefeat, len(onefeat))  # Create sample diagram
                h1 = self.elu(self.layer_normal(getattr(self, "conv{}_1".format(0))(tmpg, onefeat)))
                h1 = self.dropout(h1)
                # realtion
                h = self.elu(self.layer_normal(getattr(self, "conv{}".format(0))(g, ndata)))
                h = self.dropout(h)
                g.ndata["tmp"] = h
                readout = dgl.readout_nodes(g, "tmp", op="max")
                for index, layer in enumerate(range(self.nlayer)[1:-1]):
                    # node
                    h1 = self.elu(self.layer_normal(getattr(self, "conv{}_1".format(layer))(tmpg, h1 + readout)))
                    h1 = self.dropout(h1)
                    # relation
                    h = self.elu(self.layer_normal(getattr(self, "conv{}".format(layer))(g, h)))
                    h = self.dropout(h)
                    g.ndata["tmp"] = h
                    readout = dgl.readout_nodes(g, "tmp", op="max")
                # node
                h1 = getattr(self, "conv{}_1".format(self.nlayer - 1))(tmpg, h1 + readout)
                # relation
                h = getattr(self, "conv{}".format(self.nlayer - 1))(g, h)
                g.ndata["h"] = h
                h = dgl.readout_nodes(g, "h", op="max")
                h = self.weight * h + (1 - self.weight) * h1
                h = torch.softmax(h, dim=-1)
            return h
    class GNN_complex2_explainnode(nn.Module):
        def __init__(self):
            super(GNN_complex2_explainnode, self).__init__()
            self.nlayer = net.nlayer
            # relation
            self.conv0=net.conv0
            for layer in range(net.nlayer)[1:-1]:
                setattr(self, "conv{}".format(layer), getattr(net,"conv{}".format(layer)))
            setattr(self, "conv{}".format(net.nlayer - 1), getattr(net,"conv{}".format(net.nlayer - 1)))
            #node
            setattr(self, "conv{}_1".format(0),getattr(net,"conv{}_1".format(0)))
            for layer in range(net.nlayer)[1:-1]:
                setattr(self, "conv{}_1".format(layer),getattr(net,"conv{}_1".format(layer)))
            setattr(self, "conv{}_1".format(net.nlayer - 1),getattr(net,"conv{}_1".format(net.nlayer - 1)))
            self.elu = net.elu
            self.modenode = net.modenode
            self.dropout = net.dropout
            self.weight =net.weight

        def layer_normal(self, x, dim=None):
            if dim is None:
                std, mean = torch.std_mean(x)
            else:
                std, mean = torch.std_mean(x, dim, keepdim=True)
            return (x - mean) / (std + 1e-05)

        def forward(self, graph, feat,relationg,relationfeat,eweight=None):  #Normal graph classification and terminal node classification by adding terminal nodes
            tmpg,onefeat=graph,feat
            g,ndata=relationg,relationfeat
            with g.local_scope():#and tmp.local_scope易导致runtime error: device-side triggered
                # node
                g.ndata["feat"] = ndata
                h1 = self.elu(self.layer_normal(getattr(self, "conv{}_1".format(0))(tmpg, onefeat)))
                h1 = self.dropout(h1)
                # realtion
                h = self.elu(self.layer_normal(getattr(self, "conv{}".format(0))(g, ndata)))
                h = self.dropout(h)
                g.ndata["tmp"] = h
                readout = dgl.readout_nodes(g, "tmp", op="max")
                for index, layer in enumerate(range(self.nlayer)[1:-1]):
                    # node
                    h1 = self.elu(self.layer_normal(getattr(self, "conv{}_1".format(layer))(tmpg, h1 + readout)))
                    h1 = self.dropout(h1)
                    # relation
                    h = self.elu(self.layer_normal(getattr(self, "conv{}".format(layer))(g, h)))
                    h = self.dropout(h)
                    g.ndata["tmp"] = h
                    readout = dgl.readout_nodes(g, "tmp", op="max")
                # node
                h1 = getattr(self, "conv{}_1".format(self.nlayer - 1))(tmpg, h1 + readout)
                # relation
                h = getattr(self, "conv{}".format(self.nlayer - 1))(g, h)
                g.ndata["h"] = h
                h = dgl.readout_nodes(g, "h", op="max")
                h = self.weight * h + (1 - self.weight) * h1
                h = torch.softmax(h, dim=-1)
            return h
    print(GNN_complex2_explaing())
    print(GNN_complex2_explainnode())
    from dgl.nn.pytorch.explain import GNNExplainer
    import shap
    gnnexplainer=GNNExplainer(GNN_complex2_explaing(),num_hops=1)
    # explainer=shap.DeepExplainer(GNN_complex2_explaing(),data=[typicalsampleg,typicalsampleg.ndata["feat"][:,selecti]])
    # explainer.shap_values([typicalsampleg,typicalsampleg.ndata["feat"][:,selecti]]);exit(0)
    print(typicalsampleg.is_homogeneous)
    featmask,edgemask=gnnexplainer.explain_graph(graph=typicalsampleg,feat=typicalsampleg.ndata["feat"][:,selecti])
    print(featmask)
    print(edgemask.shape)
    # 保存到图
    typicalsampleg.edata["edgemask"] = edgemask
    # importance node and edge
    tmpig = []
    for g in dgl.unbatch(typicalsampleg):
        import dgl.function as fn
        g.update_all(fn.copy_edge("edgemask", "tmp"), fn.sum("tmp", "sumedgemask"))
        zeros=torch.zeros([g.num_nodes()], device=featmask.device)
        nidmask = torch.ones([g.num_nodes()], device=featmask.device)
        print("g.num_nodes: ",g.num_nodes())
        if g.num_nodes() > 10:
            notimportant_nid = torch.topk(g.ndata["sumedgemask"], g.num_nodes() - 10, largest=False)[1]
            nidmask[notimportant_nid] =zeros[notimportant_nid]
            tmpgg = dgl.remove_nodes(g, notimportant_nid, store_ids=True)
            importantnode_eid = tmpgg.edata[dgl.EID]  # Edges of important nodes
        else:
            tmpgg=g
            importantnode_eid = tmpgg.edges(form="eid")  # Edges of important nodes
        g.ndata["importantnode_mask"] = nidmask
        tu, tv = g.edges()
        loopmask_t = (tu == tv).float()  # Calculate spin edge mask
        select_relativeeid = torch.topk((g.edata["edgemask"]*(1-loopmask_t))[importantnode_eid], 15)[1]
        importnode_importedge_eid = importantnode_eid[select_relativeeid]
        g.edata["importantnode_importantedge"] = torch.zeros([g.num_edges()], device=select_relativeeid.device)
        g.edata["importantnode_importantedge"][importnode_importedge_eid] = g.edata["edgemask"][
            importnode_importedge_eid]
        tmpig.append(g)
    typicalsampleg = dgl.batch(tmpig)
    #保存到表
    edgetable_g=torch.stack(typicalsampleg.edges(),dim=1)#边表
    nodetable_g=torch.cat([typicalsampleg.ndata["feat"][:,selecti],typicalsampleg.ndata["pixelalbel"],typicalsampleg.ndata["importantnode_mask"].flatten().unsqueeze(dim=1),typicalsampleg.ndata["sumedgemask"].flatten().unsqueeze(dim=1)],dim=1)
    edgetable_g=torch.cat([edgetable_g,edgemask.unsqueeze(dim=1),typicalsampleg.edata["importantnode_importantedge"].flatten().unsqueeze(dim=1)],dim=1)#explaingraph,Add to list
    featmaskadd=torch.zeros([len(featmask)+3],device=featmask.device)
    featmaskadd[:-3]=featmask
    nodetable_g=torch.cat([nodetable_g,featmaskadd.unsqueeze(dim=0)],dim=0)#explaingraph，Add line
    print(featmask.shape)
    nodecolums_g=listbyindice(columns, selecti)+["pixelalabel","importantmask","sumedgemask"]
    nodeindexs_g=[str(i) for i in range(typicalsampleg.num_nodes())]+["featmask"]
    edge_g_df=pd.DataFrame(edgetable_g.cpu().numpy(),columns=["u","v","edgemask","importantnode_importantedge"]).to_excel("tmpdir/typicaledgedf_g.xls")
    node_g_df=pd.DataFrame(nodetable_g.cpu().numpy(),index=nodeindexs_g,columns=nodecolums_g).to_excel("tmpdir/typicalnodedf_g.xls")

    #Save by slope unit
    tmpgs=[]
    for g in dgl.unbatch(typicalsampleg):
        if g.num_edges()>30:#Do not remove those with small edges
            selecteid=torch.topk(g.edata["edgemask"],g.num_edges()-30,largest=False)[1]#remove what needs to be removed
            g=dgl.remove_edges(g,selecteid)
        tmpgs.append(g)
    subedge_g=dgl.batch(tmpgs)
    subedgetable_g = torch.cat([torch.stack(subedge_g.edges(), dim=1),subedge_g.edata["edgemask"].flatten().unsqueeze(dim=1)],dim=1)  # edge table
    subedge_g_df = pd.DataFrame(subedgetable_g.cpu().numpy(), columns=["u", "v", "edgemask"]).to_excel(
        "tmpdir/typical_subedgedf_g.xls")

    #Enter graph, ndata, convert to onefeat and build the graph
    typicalsampleg.ndata["selectfeat"]=typicalsampleg.ndata["feat"][:, selecti]
    onefeat = dgl.readout_nodes(typicalsampleg, "selectfeat", op="mean")
    tmpg = dgl.knn_graph(onefeat, len(onefeat))  # Create sample diagram
    print("onefeat.device: ",onefeat.device)
    edgetabel_node=torch.stack(tmpg.edges(),dim=1)#edge table
    nodetabel_node=torch.zeros([len(onefeat),len(onefeat[0])*2+1+4+tmpg.num_nodes()*2],device=onefeat.device)#feat,slopunit label,featmask,sumedgemask,nodemask
    modefeat=torch.stack(listbyindice(onefeats,explainindex),dim=0)[:,selecti]
    print("modefeat :",modefeat.shape)
    nodetabel_node[:,:len(modefeat[0])]=modefeat
    nodetabel_node[:,len(modefeat[0])]=relationresult
    nodetabel_node[:,len(modefeat[0])+1]=noderesult
    nodetabel_node[:,len(modefeat[0])+2]=pred_[:,1]
    nodetabel_node[:, len(modefeat[0]) +3] = pred2[:, 1]
    tmpedgemasks=[];processedgemasks=[];subarrays=[];importantedgemasks=[]
    for index,nodeid in enumerate(range(tmpg.num_nodes())):
        print("nodeid :",nodeid)
        print("index :",index)
        nodeexplainer = GNNExplainer(GNN_complex2_explainnode(), num_hops=1)
        newid,sg,featmask,edgemask= nodeexplainer.explain_node(node_id=nodeid,graph=tmpg, feat=onefeat,relationg=typicalsampleg,relationfeat=typicalsampleg.ndata["feat"][:, selecti])
        print(featmask)
        tmpedgetabel=torch.zeros([tmpg.num_edges()],device=newid.device)
        tmpedgetabel[sg.edata[dgl.EID]]=edgemask
        tmpedgemasks.append(tmpedgetabel)
        #topk and 1hop transformed to edgemask
        ismask=torch.ones_like(tmpedgetabel,device=edgemask.device)#1
        onehopmask=torch.zeros_like(tmpedgetabel,device=edgemask.device)
        onehopmask[tmpg.in_edges(nodeid,form="eid")]=ismask[tmpg.in_edges(nodeid,form="eid")]
        topkmask=torch.zeros_like(tmpedgetabel,device=edgemask.device)
        topkmask[edgemask.topk(tmpg.num_nodes()-1)[1]]=ismask[edgemask.topk(tmpg.num_nodes()-1)[1]]
        processedgemasks.append(tmpedgetabel*(onehopmask+topkmask).bool().float())
        processeid=torch.where((onehopmask+topkmask).bool())[0]
        subedgearray=torch.cat([edgetabel_node[processeid],tmpedgetabel[processeid].unsqueeze(dim=1)],dim=1)
        subarrays.append(subedgearray)

        #save to graph
        tmpg.edata["edgemask_for_node_{}".format(nodeid)]=tmpedgetabel
        import dgl.function as fn
        tmpg.update_all(fn.copy_edge("edgemask_for_node_{}".format(nodeid),"tmp"),fn.sum("tmp","sumedgemask{}".format(nodeid)))
        selfnodemask=torch.zeros([tmpg.num_nodes()],device=tmpedgetabel.device)
        selfnodemask[nodeid]=1
        importantnid=torch.topk(tmpg.ndata["sumedgemask{}".format(nodeid)]*(1-selfnodemask),6)[1]
        ones=torch.ones([tmpg.num_nodes()],device=importantnid.device)
        importantnidmask=torch.zeros([tmpg.num_nodes()],device=importantnid.device)#Important node mask
        importantnidmask[importantnid]=ones[importantnid]
        tmpg.ndata["nodemask{}".format(nodeid)]=importantnidmask#01
        u,v,ineid=tmpg.in_edges(nodeid,form="all")
        important_ineids=[]#Edge index involving important nodes
        print("important node:",importantnid)
        for ii,tmpnid in enumerate(importantnid.tolist()):#Traverse important nodes, #Do not use the same local variable index, it is easy to go wrong
            uindex=torch.where(u==tmpnid)[0]
            if tmpnid not in v.tolist():#In order to display better, remove the spin edges
                important_ineids.append(ineid[uindex].item())
        tmpg.edata["importedgemask{}".format(nodeid)]=torch.zeros([tmpg.num_edges()],device=ineid.device)
        tmpg.edata["importedgemask{}".format(nodeid)][important_ineids]=tmpg.edata["edgemask_for_node_{}".format(nodeid)][important_ineids]

        print(labels[explainindex[nodeid]])
        nodetabel_node[index, len(modefeat[0]+4)] = labels[explainindex[nodeid]]
        nodetabel_node[index, len(modefeat[0]) +5:2*len(modefeat[0])+5] = featmask
        nodetabel_node[:,2*len(modefeat[0])+5+index:2*len(modefeat[0])+5+index+1]=tmpg.ndata["sumedgemask{}".format(nodeid)].reshape(-1,1)
        nodetabel_node[:,2*len(modefeat[0])+5+tmpg.num_nodes()+index:2*len(modefeat[0])+5+tmpg.num_nodes()+index+1]=importantnidmask.reshape(-1,1)#01
        importantedgemasks.append(tmpg.edata["importedgemask{}".format(nodeid)])
        print("{} node finish".format(nodeid))
    edgetabel_node=torch.cat([edgetabel_node,torch.stack(tmpedgemasks,dim=1),torch.stack(processedgemasks,dim=1),torch.stack(importantedgemasks,dim=1)],dim=1)
    nodecolumns_node=["modefeat "+sen for sen in listbyindice(columns, selecti)]+["slopunit label"]+["mask "+sen for sen in listbyindice(columns, selecti)]+["sumedgemask{}".format(i) for i in range(tmpg.num_nodes())]+["improtantnodemask{}".format(i) for i in range(tmpg.num_nodes())]
    edgecolumns_node=["u","v"]+["edgemask for node {}".format(i) for i in range(tmpg.num_nodes())]+["postedgemask{}".format(i) for i in range(tmpg.num_nodes())]+["improtantedgemask{}".format(i) for i in range(tmpg.num_nodes())]
    nodedf_node=pd.DataFrame(nodetabel_node.cpu().numpy(),columns=nodecolumns_node).to_excel("tmpdir/typicalnodedf_node.xls")
    edgedf_node=pd.DataFrame(edgetabel_node.cpu().numpy(),columns=edgecolumns_node).to_excel("tmpdir/typicaledgedf_node.xls")
    for index,subedge in enumerate(subarrays):
        pd.DataFrame(subedge.cpu().numpy(),columns=["u","v","edgemask"]).to_excel("tmpdir/typicalsample{}.xls".format(index))
    #save to graph
    tmpg.ndata["slopunit_label"]=torch.stack(listbyindice(labels,explainindex),dim=0).flatten().to(device)
    tmpg.ndata["feat"]=onefeat
    print("slopunit label: ",torch.stack(listbyindice(labels,explainindex),dim=0).flatten().shape)
    dgl.save_graphs("tmpdir/typical_gnode.g",[typicalsampleg.cpu(),tmpg.cpu()])


